import os
from openpyxl import load_workbook
import xlrd
from ingest.fetch.raw import RawPort
from ingest.fetch.metadata import BaseMetadata
from ingest.errors import RawLoadError, MetadataUnavailableError

class ExcelAdapter(RawPort):

    def fetch_raw(self, path: str) -> list[dict]:
        try:
            ext = os.path.splitext(path)[1].lower()

            if ext == ".xlsx":
                return self._read_xlsx(path)
            elif ext == ".xls":
                return self._read_xls(path)
            else:
                raise RawLoadError(f"Unsupported Excel format: {ext}")

        except Exception as e:
            raise RawLoadError(f"ExcelAdapter fetchRaw error for '{path}': {e}")

    def _read_xlsx(self, path: str) -> list[dict]:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        return self._normalize_rows(rows)

    def _read_xls(self, path: str) -> list[dict]:
        book = xlrd.open_workbook(path)
        sheet = book.sheet_by_index(0)
        rows = []

        for r in range(sheet.nrows):
            row = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
            rows.append(row)

        return self._normalize_rows(rows)

    def _normalize_rows(self, rows: list[list]) -> list[dict]:
        if not rows:
            return []

        def count_non_empty(row):
            return sum(1 for c in row if c not in (None, ""))

        non_empty_counts = [count_non_empty(r) for r in rows]
        header_index = non_empty_counts.index(max(non_empty_counts))
        header_row = rows[header_index]
        max_cols = max(len(r) for r in rows)

        header = []
        for i in range(max_cols):
            if i < len(header_row) and header_row[i] not in (None, ""):
                header.append(str(header_row[i]).strip())
            else:
                header.append(f"col_{i+1}")

        records = []
        for idx, r in enumerate(rows):
            if idx == header_index:
                continue

            normalized = list(r) + [None] * (max_cols - len(r))
            record = {col: val for col, val in zip(header, normalized)}
            records.append(record)
        return records

    def fetch_metadata(self, path: str) -> BaseMetadata | None:
        raise MetadataUnavailableError(f"Excel format does not support metadata")

