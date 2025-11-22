import os
from openpyxl import load_workbook
import xlrd
from ingest.fetch.raw import RawPort


class TabAdapter(RawPort):

    def fetchRaw(self, resource: str, params=None) -> list[dict]:
        ext = os.path.splitext(resource)[1].lower()
        if ext == ".xlsx":
            return self._read_xlsx(resource)
        elif ext == ".xls":
            return self._read_xls(resource)
        else:
            raise ValueError(f"Unsupported tabular format: {ext}")

    def _read_xlsx(self, resource: str) -> list[dict]:
        wb = load_workbook(resource, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        return self._normalize_rows(rows)

    def _read_xls(self, resource: str) -> list[dict]:
        book = xlrd.open_workbook(resource)
        sheet = book.sheet_by_index(0)
        rows = []

        for r in range(sheet.nrows):
            row_values = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
            rows.append(row_values)

        return self._normalize_rows(rows)

    def _normalize_rows(self, rows: list[list]) -> list[dict]:
        if not rows:
            return []

        def count_non_empty(row):
            return sum(1 for c in row if c not in (None, ""))

        non_empty_counts = [count_non_empty(r) for r in rows]
        max_non_empty = max(non_empty_counts)

        header_index = non_empty_counts.index(max_non_empty)
        header_row = rows[header_index]
        max_cols = max(len(r) for r in rows)

        header = []
        for i in range(max_cols):
            if i < len(header_row) and header_row[i] not in (None, ""):
                header.append(str(header_row[i]).strip())
            else:
                header.append(f"col_{i+1}")

        records = []
        for idx, r in enumerate(rows):
            if idx == header_index:
                continue

            normalized = list(r) + [None] * (max_cols - len(r))
            record = {col_name: val for col_name, val in zip(header, normalized)}
            records.append(record)
        return records
